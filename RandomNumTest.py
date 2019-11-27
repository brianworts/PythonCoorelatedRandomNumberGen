import numpy as np
from scipy import stats

import openpyxl as xl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook


from scipy.linalg import cholesky

import pandas as pd
import matplotlib.pyplot as plt

import xlsxwriter

import sys
#np.set_printoptions(threshold=sys.maxsize)

def getShockVals(prices,correlationMatrix,volatility,numIters):
    allShockValues = np.zeros((numIters, len(prices))) #Create a 1000 x 72 matrix of 0's

    choleskyMatrix = cholesky(correlationMatrix)
    i = 0
    for i in range(numIters):
        rand = np.random.lognormal(0, volatility) #Create 1 x 72 matrix of random variables from stdev's
        shockValue = np.dot(rand, choleskyMatrix) / 3.32372 #Average of shock values is 3.32372???
        allShockValues[i,:] = shockValue

    #print('allShockValues' + str(allShockValues))
    print('ShockValueMean: ' + str(np.mean(allShockValues)))

    return allShockValues


path = 'Path to correlationMatrix excel file'

wb = xl.load_workbook(filename=path)
ws = wb['Sheet1']

prices = np.ravel(pd.read_excel(path, 'Sheet1', index_col =  0, usecols = 'B:D', skiprows = 1).dropna(),order = 'C') #36x2 to 72x 1 #Order C goes left to right then top to bottom, F would do top to bottom the left to right
correlationMatrix = pd.read_excel(path, 'Sheet1', index_col =  0, usecols = 'AF:CZ', skiprows = 2) #72x72
volatility = np.ravel(pd.read_excel(path, 'Sheet1', usecols = 'P:Q', skiprows = 1).dropna(),order = 'C') #np.array([0.2,0.25]),   36x2 to 72x 1
numIters = 1000

print('prices: ' + str(prices))
print('correlationMatrix: ' + str(correlationMatrix))
print('volatility: ' + str(volatility))

allShockValues = getShockVals(prices, correlationMatrix, volatility, numIters)

wsPY = wb['PythonShockVals']
i = 0
j = 0
print('inserting to Excel...')
for i in range(numIters): #Iterate 1000 times (top to bottom)
    for j in range(len(prices)): #Iterate 72 times (left to right)
        insertcolumn = get_column_letter(j+1) #Excel rows and columns begin at 1, while excel 2d arrays begin at 0
        insertCell = insertcolumn + str(i+1)
        wsPY[insertCell] = allShockValues[i,j]

wb.save(path)
print('Inserted')
print('Graphing...')

allShockValuesFlat = np.ravel(allShockValues, order = 'C')

plt.hist(allShockValuesFlat, bins = 25, log = False, label = 'allShockValues')
plt.legend()
